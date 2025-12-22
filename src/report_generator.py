"""
Report generation functions - Excel reports with charts.
"""
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Add parent directory to path for imports
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import OUTPUT_DIR, REPORT_TITLE

logger = logging.getLogger(__name__)


def create_excel_report(
    df: pd.DataFrame,
    analysis_results: dict,
    output_path: Path = None
) -> Path:
    """
    Create a formatted Excel report with multiple sheets and charts.
    
    Args:
        df: Full ticket DataFrame
        analysis_results: Dictionary containing analysis DataFrames
        output_path: Where to save the report
        
    Returns:
        Path to the generated report
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"ticket_report_{timestamp}.xlsx"
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    def style_header_row(ws, num_cols):
        """Apply styling to header row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def write_dataframe(ws, df, start_row=1):
        """Write DataFrame to worksheet with formatting."""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == start_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
    
    # === Summary Sheet ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_stats = analysis_results.get("summary_stats", {})
    ws_summary["A1"] = REPORT_TITLE
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    row = 4
    for key, value in summary_stats.items():
        ws_summary.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws_summary.cell(row=row, column=2, value=value)
        row += 1
    
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20
    
    # === Tickets by Category Sheet ===
    if "by_category" in analysis_results:
        ws_category = wb.create_sheet("By Category")
        write_dataframe(ws_category, analysis_results["by_category"])
        
        # Add bar chart
        chart = BarChart()
        chart.title = "Tickets by Category"
        chart.x_axis.title = "Category"
        chart.y_axis.title = "Count"
        
        data = Reference(ws_category, min_col=2, min_row=1, 
                        max_row=len(analysis_results["by_category"]) + 1)
        cats = Reference(ws_category, min_col=1, min_row=2, 
                        max_row=len(analysis_results["by_category"]) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_category.add_chart(chart, "E2")
        
        for col in ["A", "B", "C"]:
            ws_category.column_dimensions[col].width = 15
    
    # === Resolution Time Sheet ===
    if "resolution_by_priority" in analysis_results:
        ws_resolution = wb.create_sheet("Resolution Time")
        write_dataframe(ws_resolution, analysis_results["resolution_by_priority"])
        
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws_resolution.column_dimensions[col].width = 18
    
    # === Team Performance Sheet ===
    if "team_performance" in analysis_results:
        ws_team = wb.create_sheet("Team Performance")
        write_dataframe(ws_team, analysis_results["team_performance"])
        
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_team.column_dimensions[col].width = 20
    
    # === Trend Sheet ===
    if "trend_daily" in analysis_results:
        ws_trend = wb.create_sheet("Trends")
        trend_df = analysis_results["trend_daily"]
        write_dataframe(ws_trend, trend_df)
        
        # Add line chart
        chart = LineChart()
        chart.title = "Ticket Volume Trend"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Tickets"
        chart.style = 10
        
        data = Reference(ws_trend, min_col=2, min_row=1, max_row=len(trend_df) + 1)
        chart.add_data(data, titles_from_data=True)
        ws_trend.add_chart(chart, "D2")
        
        ws_trend.column_dimensions["A"].width = 15
        ws_trend.column_dimensions["B"].width = 15
    
    # === SLA Compliance Sheet ===
    if "sla_compliance" in analysis_results:
        ws_sla = wb.create_sheet("SLA Compliance")
        write_dataframe(ws_sla, analysis_results["sla_compliance"])
        
        for col in ["A", "B", "C", "D", "E"]:
            ws_sla.column_dimensions[col].width = 20
    
    # === Raw Data Sheet ===
    ws_raw = wb.create_sheet("Raw Data")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_raw.cell(row=r_idx, column=c_idx, value=value)
    style_header_row(ws_raw, len(df.columns))
    
    # Save workbook
    wb.save(output_path)
    logger.info(f"Report saved to: {output_path}")
    
    return output_path


def create_charts(df: pd.DataFrame, output_dir: Path = None) -> list:
    """
    Create standalone chart images.
    
    Args:
        df: Ticket DataFrame
        output_dir: Directory to save charts
        
    Returns:
        List of paths to generated chart files
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    output_dir.mkdir(parents=True, exist_ok=True)
    chart_files = []
    
    # Set style
    plt.style.use("seaborn-v0_8-whitegrid")
    
    # Tickets by Category - Bar Chart
    fig, ax = plt.subplots(figsize=(10, 6))
    category_counts = df["category"].value_counts()
    category_counts.plot(kind="bar", ax=ax, color="steelblue")
    ax.set_title("Tickets by Category", fontsize=14, fontweight="bold")
    ax.set_xlabel("Category")
    ax.set_ylabel("Count")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    
    chart_path = output_dir / "chart_category.png"
    fig.savefig(chart_path, dpi=150)
    plt.close(fig)
    chart_files.append(chart_path)
    
    # Daily Trend - Line Chart
    fig, ax = plt.subplots(figsize=(12, 5))
    daily = df.groupby(df["created_date"].dt.date).size()
    daily.plot(kind="line", ax=ax, color="steelblue", marker="o", markersize=3)
    ax.set_title("Daily Ticket Volume", fontsize=14, fontweight="bold")
    ax.set_xlabel("Date")
    ax.set_ylabel("Tickets")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    
    chart_path = output_dir / "chart_trend.png"
    fig.savefig(chart_path, dpi=150)
    plt.close(fig)
    chart_files.append(chart_path)
    
    # Priority Distribution - Pie Chart
    fig, ax = plt.subplots(figsize=(8, 8))
    priority_counts = df["priority"].value_counts()
    colors = ["#2ecc71", "#f1c40f", "#e67e22", "#e74c3c"]
    priority_counts.plot(kind="pie", ax=ax, autopct="%1.1f%%", colors=colors)
    ax.set_title("Priority Distribution", fontsize=14, fontweight="bold")
    ax.set_ylabel("")
    plt.tight_layout()
    
    chart_path = output_dir / "chart_priority.png"
    fig.savefig(chart_path, dpi=150)
    plt.close(fig)
    chart_files.append(chart_path)
    
    logger.info(f"Generated {len(chart_files)} charts")
    return chart_files


if __name__ == "__main__":
    # Test report generation
    logging.basicConfig(level=logging.INFO)
    
    from data_loader import load_tickets
    from analysis import (
        tickets_by_category,
        avg_resolution_time_by_priority,
        team_performance,
        tickets_over_time,
        sla_compliance,
        generate_summary_stats
    )
    
    try:
        df = load_tickets()
        
        analysis_results = {
            "summary_stats": generate_summary_stats(df),
            "by_category": tickets_by_category(df),
            "resolution_by_priority": avg_resolution_time_by_priority(df),
            "team_performance": team_performance(df),
            "trend_daily": tickets_over_time(df, "D"),
            "sla_compliance": sla_compliance(df)
        }
        
        report_path = create_excel_report(df, analysis_results)
        print(f"Report generated: {report_path}")
        
        chart_files = create_charts(df)
        print(f"Charts generated: {chart_files}")
        
    except FileNotFoundError:
        print("Run generate_mock_data.py first to create test data.")
